"""
S-245 (Create Employee Profile) + S-246 (Mark Employee as Bench) +
S-247 (View Bench Pool) + S-248 (Bench Duration & Aging Report) — API
Endpoints
=========================================================================
Prefix: /employees
Tag:    employees

No employee REST API of any kind previously existed in this codebase
despite the Employee model, bench_pool lifecycle, and bench_periods
history all being real, tested backend (this and earlier Phase 4
rounds). This closes that gap for the foundational pieces four stories
in the EPIC-05/Resource & Bench Management cluster need.

Auth: same posture as every Phase 4 endpoint this program
(get_current_hr_or_admin -- any internal user). No employee-specific
RBAC permission exists yet.

Routes (static paths registered before the /{employee_id} catch-all,
so "bench-pool"/"bench-aging-alerts" never get swallowed as an id):
  POST   /employees
  GET    /employees
  GET    /employees/bench-pool
  GET    /employees/bench-aging-alerts
  GET    /employees/{employee_id}
  POST   /employees/{employee_id}/mark-bench
  POST   /employees/{employee_id}/remove-from-bench
  GET    /employees/{employee_id}/bench-history
"""
import io
import json
from datetime import date, datetime

from fastapi import APIRouter, Depends, HTTPException, UploadFile
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.dependencies import get_current_hr_or_admin
from app.models.candidate import Candidate
from app.models.employee import Employee, EmployeeEngineHistory
from app.models.resource_management import BenchPoolEntry
from app.models.user import Users
from app.schemas.employee import (
    BenchAgingAlertItem,
    BenchAgingAlertsResponse,
    BenchCostSummaryItem,
    BenchCostSummaryResponse,
    BenchPeriodHistoryResponse,
    BenchPeriodItem,
    BulkImportResponse,
    BulkImportRowError,
    ConvertCandidateRequest,
    EmployeeCreateRequest,
    EmployeeItem,
    EmployeeListResponse,
    EngineHistoryItem,
    EngineHistoryResponse,
    MarkBenchRequest,
    RecordUtilizationRequest,
    StaffingEligibilityResponse,
    UtilizationHistoryItem,
    UtilizationHistoryResponse,
    UtilizationSummaryItem,
    UtilizationSummaryResponse,
)
from app.services.employee_service import (
    DuplicateEmployeeEmail,
    convert_candidate_to_employee,
    create_employee_profile,
    generate_employee_number,
)
from app.services.resource_management_service import (
    LOW_UTILIZATION_THRESHOLD_PCT,
    check_bench_aging_alerts,
    get_bench_duration_days,
    get_bench_period_history,
    get_current_bench_pool,
    get_latest_utilization_by_employee,
    get_utilization_history,
    is_staffing_eligible,
    mark_employee_on_bench,
    record_weekly_utilization_metric,
    remove_employee_from_bench,
)

router = APIRouter(prefix="/employees", tags=["employees"])


def _skills_list(raw):
    if not raw:
        return []
    try:
        parsed = json.loads(raw)
    except (json.JSONDecodeError, TypeError):
        return []
    return parsed if isinstance(parsed, list) else []


def _to_item(db: Session, employee: Employee) -> EmployeeItem:
    bench_entry = db.query(BenchPoolEntry).filter(BenchPoolEntry.employee_id == employee.id).first()
    return EmployeeItem(
        id=employee.id,
        employee_number=employee.employee_number,
        first_name=employee.first_name,
        last_name=employee.last_name,
        email=employee.email,
        status=employee.status,
        employment_type=employee.employment_type,
        current_title=employee.current_title,
        current_skills=_skills_list(employee.current_skills),
        work_location=employee.work_location,
        delivery_engine=employee.delivery_engine,
        engine_entry_date=employee.engine_entry_date,
        core_certified=employee.core_certified,
        core_certified_date=employee.core_certified_date,
        joining_date=employee.joining_date,
        base_salary_usd_cents=employee.base_salary_usd_cents,
        billing_rate_usd_cents=employee.billing_rate_usd_cents,
        is_on_bench=bench_entry is not None,
        bench_days=get_bench_duration_days(bench_entry) if bench_entry else None,
    )


def _get_employee_or_404(db: Session, employee_id: str) -> Employee:
    employee = db.query(Employee).filter(Employee.id == employee_id).first()
    if employee is None:
        raise HTTPException(status_code=404, detail="Employee not found.")
    return employee


@router.post("", response_model=EmployeeItem, summary="Create an employee profile")
def create_employee(
    body: EmployeeCreateRequest,
    db: Session = Depends(get_db),
    current_user: Users = Depends(get_current_hr_or_admin),
):
    fields = {}
    if body.current_title is not None:
        fields["current_title"] = body.current_title
    if body.current_skills is not None:
        fields["current_skills"] = json.dumps(body.current_skills)
    if body.employment_type is not None:
        fields["employment_type"] = body.employment_type
    if body.work_location is not None:
        fields["work_location"] = body.work_location
    if body.base_salary_usd_cents is not None:
        fields["base_salary_usd_cents"] = body.base_salary_usd_cents
    if body.billing_rate_usd_cents is not None:
        fields["billing_rate_usd_cents"] = body.billing_rate_usd_cents
    if body.nationality is not None:
        fields["nationality"] = body.nationality

    try:
        employee = create_employee_profile(
            db, tenant_id=current_user.tenant_id,
            first_name=body.first_name, last_name=body.last_name, email=body.email,
            joining_date=body.joining_date, changed_by=current_user.UserID,
            **fields,
        )
    except DuplicateEmployeeEmail as exc:
        raise HTTPException(status_code=409, detail=str(exc))
    db.commit()
    db.refresh(employee)
    return _to_item(db, employee)


@router.post(
    "/convert-candidate/{candidate_id}", response_model=EmployeeItem,
    summary="Convert a candidate to an employee (HRMS-0708 minimal slice)",
)
def convert_candidate(
    candidate_id: str,
    body: ConvertCandidateRequest,
    db: Session = Depends(get_db),
    current_user: Users = Depends(get_current_hr_or_admin),
):
    candidate = db.query(Candidate).filter(Candidate.candidateID == candidate_id).first()
    if candidate is None:
        raise HTTPException(status_code=404, detail="Candidate not found.")
    if db.query(Employee).filter(Employee.candidate_id == candidate_id).first():
        raise HTTPException(status_code=409, detail=f"Candidate {candidate_id} has already been converted to an employee.")

    fields = {"employee_number": generate_employee_number(db, current_user.tenant_id, "BLX")}
    if body.current_title is not None:
        fields["current_title"] = body.current_title
    if body.current_skills is not None:
        fields["current_skills"] = json.dumps(body.current_skills)
    if body.employment_type is not None:
        fields["employment_type"] = body.employment_type
    if body.work_location is not None:
        fields["work_location"] = body.work_location
    if body.base_salary_usd_cents is not None:
        fields["base_salary_usd_cents"] = body.base_salary_usd_cents
    if body.billing_rate_usd_cents is not None:
        fields["billing_rate_usd_cents"] = body.billing_rate_usd_cents
    if body.nationality is not None:
        fields["nationality"] = body.nationality

    employee = convert_candidate_to_employee(
        db, candidate, joining_date=body.joining_date, tenant_id=current_user.tenant_id,
        changed_by=current_user.UserID, **fields,
    )
    db.commit()
    db.refresh(employee)
    return _to_item(db, employee)


_BULK_IMPORT_COLUMNS = (
    "first_name", "last_name", "email", "joining_date", "current_title",
    "current_skills", "employment_type", "work_location",
    "base_salary_usd_cents", "billing_rate_usd_cents", "nationality",
)


def _parse_bulk_import_date(raw) -> date:
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    return datetime.strptime(str(raw).strip(), "%Y-%m-%d").date()


@router.post(
    "/bulk-import", response_model=BulkImportResponse,
    summary="One-time bulk employee load from an .xlsx file",
    description=(
        "Header row (first row, any casing) must include first_name, last_name, "
        "email, joining_date (YYYY-MM-DD) -- required -- plus optional current_title, "
        "current_skills (comma-separated), employment_type, work_location, "
        "base_salary_usd_cents, billing_rate_usd_cents, nationality. Each row is "
        "created independently -- one bad row (duplicate email, missing required "
        "field) is reported and skipped, it does not abort the rest of the file."
    ),
)
async def bulk_import_employees(
    file: UploadFile,
    db: Session = Depends(get_db),
    current_user: Users = Depends(get_current_hr_or_admin),
):
    raw = await file.read()
    try:
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=422, detail=f"Could not read the uploaded file as .xlsx: {exc}")
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=422, detail="The uploaded file has no rows.")

    header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    col_index = {name: header.index(name) for name in _BULK_IMPORT_COLUMNS if name in header}
    missing_required = [c for c in ("first_name", "last_name", "email", "joining_date") if c not in col_index]
    if missing_required:
        raise HTTPException(
            status_code=422,
            detail=f"Header row is missing required column(s): {', '.join(missing_required)}.",
        )

    def _cell(row, name):
        idx = col_index.get(name)
        if idx is None or idx >= len(row):
            return None
        value = row[idx]
        if value is None or (isinstance(value, str) and not value.strip()):
            return None
        return value

    created = 0
    skipped = 0
    errors = []

    for row_num, row in enumerate(rows[1:], start=2):
        if row is None or all(v is None for v in row):
            continue

        email = _cell(row, "email")
        try:
            first_name = _cell(row, "first_name")
            last_name = _cell(row, "last_name")
            joining_date_raw = _cell(row, "joining_date")
            if not first_name or not last_name or not email or not joining_date_raw:
                raise ValueError("first_name, last_name, email, and joining_date are all required.")

            fields = {}
            current_title = _cell(row, "current_title")
            if current_title:
                fields["current_title"] = str(current_title).strip()
            current_skills = _cell(row, "current_skills")
            if current_skills:
                fields["current_skills"] = json.dumps(
                    [s.strip() for s in str(current_skills).split(",") if s.strip()]
                )
            employment_type = _cell(row, "employment_type")
            if employment_type:
                fields["employment_type"] = str(employment_type).strip().upper()
            work_location = _cell(row, "work_location")
            if work_location:
                fields["work_location"] = str(work_location).strip().upper()
            base_salary = _cell(row, "base_salary_usd_cents")
            if base_salary is not None:
                fields["base_salary_usd_cents"] = int(base_salary)
            billing_rate = _cell(row, "billing_rate_usd_cents")
            if billing_rate is not None:
                fields["billing_rate_usd_cents"] = int(billing_rate)
            nationality = _cell(row, "nationality")
            if nationality:
                fields["nationality"] = str(nationality).strip()

            create_employee_profile(
                db, tenant_id=current_user.tenant_id,
                first_name=str(first_name).strip(), last_name=str(last_name).strip(),
                email=str(email).strip(), joining_date=_parse_bulk_import_date(joining_date_raw),
                changed_by=current_user.UserID, **fields,
            )
            db.commit()
            created += 1
        except DuplicateEmployeeEmail as exc:
            db.rollback()
            skipped += 1
            errors.append(BulkImportRowError(row=row_num, email=str(email) if email else None, reason=str(exc)))
        except Exception as exc:
            db.rollback()
            skipped += 1
            errors.append(BulkImportRowError(row=row_num, email=str(email) if email else None, reason=str(exc)))

    return BulkImportResponse(created=created, skipped=skipped, errors=errors)


@router.get("", response_model=EmployeeListResponse, summary="List employees")
def list_employees(
    db: Session = Depends(get_db),
    current_user: Users = Depends(get_current_hr_or_admin),
):
    employees = (
        db.query(Employee)
        .filter(Employee.tenant_id == current_user.tenant_id)
        .order_by(Employee.created_at.desc())
        .all()
    )
    return EmployeeListResponse(employees=[_to_item(db, e) for e in employees])


@router.get(
    "/bench-pool", response_model=EmployeeListResponse,
    summary="View the current bench pool (S-247)",
)
def view_bench_pool(
    db: Session = Depends(get_db),
    current_user: Users = Depends(get_current_hr_or_admin),
):
    entries = get_current_bench_pool(db, tenant_id=current_user.tenant_id)
    employees = []
    for entry in entries:
        employee = db.query(Employee).filter(Employee.id == entry.employee_id).first()
        if employee is not None:
            employees.append(_to_item(db, employee))
    return EmployeeListResponse(employees=employees)


@router.get(
    "/bench-aging-alerts", response_model=BenchAgingAlertsResponse,
    summary="Employees who just crossed a 30/60/90-day bench milestone (S-248)",
)
def bench_aging_alerts(
    db: Session = Depends(get_db),
    current_user: Users = Depends(get_current_hr_or_admin),
):
    alerts = check_bench_aging_alerts(db, tenant_id=current_user.tenant_id)
    items = []
    for alert in alerts:
        employee = db.query(Employee).filter(Employee.id == alert["employee_id"]).first()
        employee_name = f"{employee.first_name} {employee.last_name}".strip() if employee else "(unknown)"
        items.append(BenchAgingAlertItem(employee_name=employee_name, **alert))
    return BenchAgingAlertsResponse(alerts=items)


@router.get(
    "/utilization-summary", response_model=UtilizationSummaryResponse,
    summary="Employee Utilization Dashboard (S-254) -- latest utilization per employee + low-utilization alerts",
)
def utilization_summary(
    db: Session = Depends(get_db),
    current_user: Users = Depends(get_current_hr_or_admin),
):
    latest_by_employee = get_latest_utilization_by_employee(db, tenant_id=current_user.tenant_id)
    items = []
    pct_values = []
    low_count = 0
    for employee_id, metric in latest_by_employee.items():
        employee = db.query(Employee).filter(Employee.id == employee_id).first()
        if employee is None:
            continue
        pct = float(metric.utilization_pct)
        pct_values.append(pct)
        is_low = pct < LOW_UTILIZATION_THRESHOLD_PCT
        if is_low:
            low_count += 1
        items.append(UtilizationSummaryItem(
            employee_id=employee_id,
            employee_name=f"{employee.first_name} {employee.last_name}".strip(),
            bu_id=employee.bu_id,
            latest_utilization_pct=pct,
            latest_period_start=metric.period_start,
            is_low_utilization=is_low,
        ))
    average = round(sum(pct_values) / len(pct_values), 2) if pct_values else None
    return UtilizationSummaryResponse(
        employees=items, average_utilization_pct=average, low_utilization_count=low_count,
    )


@router.get(
    "/bench-cost-summary", response_model=BenchCostSummaryResponse,
    summary="Bench Cost Visibility (S-255) -- daily/monthly/running bench cost per employee + total",
)
def bench_cost_summary(
    db: Session = Depends(get_db),
    current_user: Users = Depends(get_current_hr_or_admin),
):
    entries = get_current_bench_pool(db, tenant_id=current_user.tenant_id)
    items = []
    total_running = 0
    for entry in entries:
        employee = db.query(Employee).filter(Employee.id == entry.employee_id).first()
        if employee is None:
            continue
        days = get_bench_duration_days(entry)
        daily = entry.bench_cost_usd_cents
        running = (daily * days) if daily is not None else None
        if running is not None:
            total_running += running
        items.append(BenchCostSummaryItem(
            employee_id=entry.employee_id,
            employee_name=f"{employee.first_name} {employee.last_name}".strip(),
            days_on_bench=days,
            daily_cost_usd_cents=daily,
            monthly_cost_usd_cents=(daily * 30) if daily is not None else None,
            running_total_usd_cents=running,
        ))
    return BenchCostSummaryResponse(employees=items, total_running_cost_usd_cents=total_running)


@router.get("/{employee_id}", response_model=EmployeeItem, summary="Get one employee")
def get_employee(
    employee_id: str,
    db: Session = Depends(get_db),
    current_user: Users = Depends(get_current_hr_or_admin),
):
    employee = _get_employee_or_404(db, employee_id)
    return _to_item(db, employee)


@router.get(
    "/{employee_id}/staffing-eligibility", response_model=StaffingEligibilityResponse,
    summary="Staffing Eligibility Engine (S-250) -- can this employee appear in ranking for a delivery engine?",
)
def staffing_eligibility(
    employee_id: str,
    delivery_engine: str,
    db: Session = Depends(get_db),
    current_user: Users = Depends(get_current_hr_or_admin),
):
    employee = _get_employee_or_404(db, employee_id)
    eligible, reason = is_staffing_eligible(employee, delivery_engine)
    return StaffingEligibilityResponse(
        employee_id=employee_id, delivery_engine=delivery_engine, eligible=eligible, reason=reason,
    )


@router.get(
    "/{employee_id}/engine-history", response_model=EngineHistoryResponse,
    summary="S-351/HRMS-0512 -- read-only Speciality/Core engine change audit trail",
)
def engine_history(
    employee_id: str,
    db: Session = Depends(get_db),
    current_user: Users = Depends(get_current_hr_or_admin),
):
    """Read-only per the source doc's own 'Not In Scope: do NOT build any
    UI bypass for delivery engine assignment' -- CORE can only ever be
    set via set_core_delivery_engine() (employee_service.py), not through
    this endpoint."""
    _get_employee_or_404(db, employee_id)
    rows = (
        db.query(EmployeeEngineHistory)
        .filter(EmployeeEngineHistory.employee_id == employee_id)
        .order_by(EmployeeEngineHistory.changed_at.asc())
        .all()
    )
    return EngineHistoryResponse(
        employee_id=employee_id,
        history=[
            EngineHistoryItem(
                id=h.id, from_engine=h.from_engine, to_engine=h.to_engine,
                changed_at=h.changed_at, changed_by=h.changed_by,
                approval_reference=h.approval_reference, reason=h.reason,
            )
            for h in rows
        ],
    )


@router.post(
    "/{employee_id}/record-utilization", response_model=UtilizationHistoryItem,
    summary="Calculate + snapshot one employee's utilization for a week (S-223/HRMS-0904) -- billable hours vs available hours",
)
def record_utilization(
    employee_id: str,
    body: RecordUtilizationRequest,
    db: Session = Depends(get_db),
    current_user: Users = Depends(get_current_hr_or_admin),
):
    employee = _get_employee_or_404(db, employee_id)
    metric = record_weekly_utilization_metric(db, employee, body.week_starting_date)
    db.commit()
    db.refresh(metric)
    return UtilizationHistoryItem(
        period_start=metric.period_start, utilization_pct=float(metric.utilization_pct),
        billable_hours=float(metric.billable_hours), bench_hours=float(metric.bench_hours),
    )


@router.get(
    "/{employee_id}/utilization-history", response_model=UtilizationHistoryResponse,
    summary="Weekly utilization history for one employee (S-254)",
)
def utilization_history(
    employee_id: str,
    db: Session = Depends(get_db),
    current_user: Users = Depends(get_current_hr_or_admin),
):
    _get_employee_or_404(db, employee_id)
    metrics = get_utilization_history(db, employee_id)
    return UtilizationHistoryResponse(
        employee_id=employee_id,
        history=[
            UtilizationHistoryItem(
                period_start=m.period_start, utilization_pct=float(m.utilization_pct),
                billable_hours=float(m.billable_hours), bench_hours=float(m.bench_hours),
            )
            for m in metrics
        ],
    )


@router.post(
    "/{employee_id}/mark-bench", response_model=EmployeeItem,
    summary="Mark an employee as on the bench (S-246)",
)
def mark_bench(
    employee_id: str,
    body: MarkBenchRequest,
    db: Session = Depends(get_db),
    current_user: Users = Depends(get_current_hr_or_admin),
):
    employee = _get_employee_or_404(db, employee_id)
    mark_employee_on_bench(db, employee, reason=body.reason)
    db.commit()
    db.refresh(employee)
    return _to_item(db, employee)


@router.post(
    "/{employee_id}/remove-from-bench", response_model=EmployeeItem,
    summary="Remove an employee from the bench",
)
def remove_from_bench(
    employee_id: str,
    db: Session = Depends(get_db),
    current_user: Users = Depends(get_current_hr_or_admin),
):
    employee = _get_employee_or_404(db, employee_id)
    remove_employee_from_bench(db, employee)
    db.commit()
    db.refresh(employee)
    return _to_item(db, employee)


@router.get(
    "/{employee_id}/bench-history", response_model=BenchPeriodHistoryResponse,
    summary="Full bench episode history for an employee",
)
def bench_history(
    employee_id: str,
    db: Session = Depends(get_db),
    current_user: Users = Depends(get_current_hr_or_admin),
):
    _get_employee_or_404(db, employee_id)
    periods = get_bench_period_history(db, employee_id)
    return BenchPeriodHistoryResponse(
        periods=[
            BenchPeriodItem(
                id=p.id, bench_start_date=p.bench_start_date, bench_end_date=p.bench_end_date,
                reason_for_bench=p.reason_for_bench, bench_cost_usd_cents=p.bench_cost_usd_cents,
            )
            for p in periods
        ]
    )
